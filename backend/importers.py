"""Turn what a broker or registrar already gives you into holdings.

Nobody types thirty-five stocks in by hand. Two routes in:

* **Any CSV/XLSX a broker exports.** Rather than making you rename columns to
  match ours, the headers are matched against the names brokers actually use
  and the guessed mapping is shown for correction before anything is
  imported. Renaming still works -- our own names are in the alias list -- it
  just is not required.
* **A CAMS/KFintech consolidated account statement.** One password-protected
  PDF holding every mutual fund folio across both registrars.

Both routes end at the same place: a preview you confirm. Nothing is written
to the portfolio until you have seen the rows.
"""
import csv
import io
import re

# What a column can mean, and the header names brokers really use for it.
# Matched after normalisation (lowercased, punctuation stripped), so
# "Avg. Cost", "avg cost" and "AVG_COST" are the same key.
COLUMN_ALIASES = {
    "identifier": [
        "symbol", "instrument", "tradingsymbol", "trading symbol", "scrip",
        "scrip name", "scrip code", "stock", "stock symbol", "ticker",
        "nse symbol", "nse code", "bse code", "isin", "isin code",
        "security id", "identifier", "folio", "folio no", "folio number",
    ],
    "name": [
        "name", "company", "company name", "security name", "instrument name",
        "stock name", "scheme name", "scheme", "particulars", "description",
    ],
    "units": [
        "qty", "quantity", "shares", "units", "holding qty", "held qty",
        "total qty", "free qty", "balance units", "closing balance",
        "closing unit balance", "no of shares", "quantity available",
    ],
    "avg_cost": [
        "avg", "avg cost", "average price", "avg price", "buy avg",
        "buy average", "average cost", "cost per unit", "avg buy price",
        "purchase nav", "avg nav", "average nav", "buy price", "rate",
    ],
    "last_price": [
        "ltp", "last price", "current price", "market price", "closing price",
        "cmp", "nav", "current nav", "last traded price", "price",
    ],
    "invested": [
        "invested", "invested value", "investment", "investment value",
        "total cost", "cost value", "buy value", "purchase value",
        "amount invested", "total cost value", "purchase cost",
        "value at cost", "inv value", "inv amt", "invested amt",
    ],
    "current_value": [
        "current value", "market value", "present value", "valuation",
        "closing value", "market val", "value",
        # Zerodha writes "Cur. val"; others abbreviate differently again.
        "cur val", "curr val", "current val", "cur value", "curr value",
        "value at market price", "mkt value", "mkt val", "current amt",
    ],
    "purchase_date": [
        "purchase date", "buy date", "date", "transaction date",
        "date of purchase", "trade date",
    ],
}

# Fields the UI offers in the mapping dropdowns, in display order.
MAPPABLE = ["identifier", "name", "units", "avg_cost", "last_price",
            "invested", "current_value", "purchase_date"]


def normalise_header(h):
    h = re.sub(r"[^a-z0-9 ]+", " ", str(h or "").lower())
    return re.sub(r"\s+", " ", h).strip()


def sniff_columns(headers):
    """Guess which of our fields each column holds.

    Exact alias matches win over partial ones, and each field is claimed at
    most once, so a sheet with both "Price" and "Last Price" does not map
    both to last_price.
    """
    norm = [normalise_header(h) for h in headers]
    mapping, taken = {}, set()
    for field in MAPPABLE:                       # exact matches first
        aliases = COLUMN_ALIASES[field]
        for i, h in enumerate(norm):
            if i in taken or not h:
                continue
            if h in aliases:
                mapping[field] = headers[i]
                taken.add(i)
                break
    for field in MAPPABLE:                       # then contains-matches
        if field in mapping:
            continue
        for i, h in enumerate(norm):
            if i in taken or not h:
                continue
            if any(a in h or h in a for a in COLUMN_ALIASES[field]
                   if len(a) > 3):
                mapping[field] = headers[i]
                taken.add(i)
                break
    return mapping


def to_number(value):
    """Parse a spreadsheet money/quantity cell. Returns None when not a number."""
    if value is None:
        return None
    s = str(value).strip()
    if not s or s in {"-", "--", "NA", "N/A", "nil"}:
        return None
    neg = s.startswith("(") and s.endswith(")")
    s = re.sub(r"[^0-9.\-]", "", s.strip("()"))
    if not s or s in {"-", ".", "-."}:
        return None
    try:
        n = float(s)
    except ValueError:
        return None
    return -n if neg else n


def read_table(data, filename=""):
    """Rows of a CSV or XLSX as (headers, list-of-dicts)."""
    if filename.lower().endswith((".xlsx", ".xlsm")):
        return _read_xlsx(data)
    text = data.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    rows = [r for r in reader if any((c or "").strip() for c in r)]
    return _table_from_rows(rows)


# How far down a sheet to look for the header. Broker exports put a client
# id, a title and a summary block above it; thirty rows covers every one
# seen so far without scanning a whole file.
MAX_HEADER_SCAN = 30


def _header_score(cells):
    """How much a row looks like the header of a holdings table.

    Width is not the signal, which is what the first version of this got
    wrong. A CSV saved out of Excel pads *every* row to the same number of
    fields, so "widest row" picked row one -- a client id -- and every
    column came back unmapped. What actually distinguishes a header is that
    its cells are words we recognise, so that is what is counted, with the
    number of filled cells only as a tie-break.
    """
    text = [str(c).strip() for c in cells]
    filled = [c for c in text if c]
    if len(filled) < 2:
        return (0, 0)
    return (len(sniff_columns(text)), len(filled))


def _pick_header(rows):
    """Index of the most header-looking row that has data under it."""
    best_idx, best = 0, (-1, -1)
    for i in range(min(len(rows), MAX_HEADER_SCAN)):
        if i + 1 >= len(rows):      # a last line is a footer, not a header
            break
        score = _header_score(rows[i])
        if score > best:
            best_idx, best = i, score
    return best_idx, best


def _table_from_rows(rows):
    """(headers, records) from raw rows, finding the header row first."""
    if not rows:
        return [], []
    idx, _ = _pick_header(rows)
    headers = [str(h).strip() for h in rows[idx]]
    out = []
    for r in rows[idx + 1:]:
        if len([c for c in r if str(c).strip()]) < 2:
            continue
        out.append({headers[i]: (r[i] if i < len(r) else "")
                    for i in range(len(headers))})
    return headers, out


def _sheet_rows(ws):
    ws.reset_dimensions()   # see _read_xlsx
    rows = [[("" if c is None else c) for c in row]
            for row in ws.iter_rows(values_only=True)]
    return [r for r in rows if any(str(c).strip() for c in r)]


def _read_xlsx(data):
    """The best table in the workbook, whichever sheet it is on.

    Two things this has to survive. Some writers declare
    <dimension ref="A1"/> whatever the sheet actually holds; openpyxl's
    read-only mode believes that and yields one empty cell, so a perfectly
    good file imported as nothing at all. reset_dimensions() makes it read
    the rows instead.

    And a broker workbook usually has several sheets -- Equity, Mutual
    Funds, Combined -- so taking the first one is a guess. Each is scored
    the same way a header row is, and the sheet with the most recognisable
    table wins.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ValueError("Reading .xlsx needs the openpyxl package; export "
                         "the sheet as CSV instead.")
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    best, best_score = ([], []), (-1, -1)
    for name in wb.sheetnames:
        rows = _sheet_rows(wb[name])
        if not rows:
            continue
        idx, score = _pick_header(rows)
        # Recognisable columns first, then how many rows sit under them: a
        # summary tab can name a column or two but carries no holdings.
        rank = (score[0], len(rows) - idx - 1)
        if rank > best_score:
            best, best_score = _table_from_rows(rows), rank
    return best


def derive_quantities(units=None, avg_cost=None, last_price=None,
                      invested=None, current_value=None):
    """Fill in whichever of units / cost / price is missing.

    Units are the quantity the app stores; invested and current value are
    products of it. But nobody reads unit counts off a screen -- what people
    can see is what a holding cost and what it is worth. So given a price,
    units follow from the value, and the cost per unit follows from the
    invested amount. Returns (units, avg_cost, last_price).

    Order matters: units first, because the per-unit figures are derived
    from them. Deriving cost per unit first and units after produced a cost
    of zero for any row that reported only money.
    """
    units = units or None
    avg_cost = avg_cost or None
    last_price = last_price or None
    invested = invested or None
    current_value = current_value or None

    if not units:
        if current_value and last_price:
            units = current_value / last_price
        elif invested and avg_cost:
            units = invested / avg_cost
    if units:
        if not avg_cost and invested:
            avg_cost = invested / units
        if not last_price and current_value:
            last_price = current_value / units
    return units, avg_cost, (last_price or avg_cost)


def build_rows(records, mapping, asset_class="stock", owner="Me"):
    """Apply a mapping to raw records and derive what is missing.

    Brokers report either an average price or a total invested value, rarely
    both, so whichever is absent is derived from the other. Rows without a
    usable quantity or name are reported rather than silently dropped.
    """
    out, skipped = [], []
    for i, rec in enumerate(records):
        def val(field):
            col = mapping.get(field)
            return rec.get(col) if col else None

        name = (str(val("name") or "").strip()
                or str(val("identifier") or "").strip())
        units = to_number(val("units"))
        avg = to_number(val("avg_cost"))
        last = to_number(val("last_price"))
        invested = to_number(val("invested"))
        current = to_number(val("current_value"))

        units, avg, last = derive_quantities(units, avg, last, invested,
                                             current)

        if not name:
            skipped.append("row %d: no name or symbol" % (i + 2))
            continue
        if not units or units <= 0:
            skipped.append(
                "row %d (%s): no quantity, and none could be worked out — "
                "a value needs a price beside it, or an invested amount "
                "needs an average cost." % (i + 2, name[:30]))
            continue

        ident = str(val("identifier") or "").strip()
        pdate = str(val("purchase_date") or "").strip()
        out.append({
            "owner": owner, "asset_class": asset_class, "name": name[:120],
            "identifier": ident[:60], "units": round(units, 4),
            "avg_cost": round(avg or 0.0, 4), "last_price": round(last or 0.0, 4),
            "invested": round((invested if invested is not None
                               else units * (avg or 0)), 2),
            "current_value": round((current if current is not None
                                    else units * (last or 0)), 2),
            "purchase_date": (pdate[:10]
                              if re.match(r"\d{4}-\d{2}-\d{2}", pdate)
                              else ""),
        })
    return out, skipped


# --------------------------------------------------------------------------
# CAMS / KFintech consolidated account statement
#
# Registrars issue two of these and people request whichever they find:
#
# * the **Consolidated Account Summary** -- one table row per folio, ending in
#   cost / units / NAV / market value;
# * the **detailed statement** -- a block per scheme carrying the folio, the
#   nominees, every transaction since inception and a closing line.
#
# Both are parsed here, anchored on the ISIN: it appears exactly once per
# holding and is the only field neither layout wraps or abbreviates.
# --------------------------------------------------------------------------

# Indian ISINs are IN + E/F/9 + 9 more. A word boundary cannot be used at
# the start: extracted text frequently glues the folio to the ISIN with no
# space ("90722941761/0INF846K01EW2"), which silently dropped those rows.
ISIN_TOKEN = re.compile(r"(?<![A-Z])IN[EF0-9][A-Z0-9]{9}(?![A-Z0-9])")
DECIMAL_TOKEN = re.compile(r"\d[\d,]*\.\d+")
NAV_DATE_TOKEN = re.compile(r"\d{2}-[A-Za-z]{3}-\d{4}")
SCHEME_CODE_PREFIX = re.compile(r"^[A-Z0-9]{2,12}\s*[-–]\s*")
REGISTRAR_TOKEN = re.compile(r"\b(CAMS|KFINTECH|KARVY)\b", re.I)


def _clean_scheme(text):
    text = re.sub(r"\s+", " ", text).strip(" -–:")
    text = SCHEME_CODE_PREFIX.sub("", text)          # drop "128TSDGG - "
    text = re.sub(r"\((?:Non[\s-]?Demat|Demat)\)", "", text, flags=re.I)
    text = re.sub(r"\(\s*Advisor\s*:[^)]*\)", "", text, flags=re.I)
    text = re.sub(r"\bISIN\b\s*:?\s*", "", text, flags=re.I)
    text = re.sub(r"\bRegistrar\b\s*:?\s*(?:CAMS|KFINTECH|KARVY)?", "", text,
                  flags=re.I)
    return re.sub(r"\s+", " ", text).strip(" -–:,")


# Statement furniture that has turned up where a scheme name should be: a
# generation timestamp, a software version banner, a page marker. A CAS
# produced "260826143717 Version:V3.5 Live-1018" as a fund name, which is
# not a fund and never will be.
FURNITURE = re.compile(
    r"\d{10,}|\bversion\s*:|\bpage\s*\d|\bstatement\s+(?:period|date)\b"
    r"|\bgenerated\s+on\b|^\W*$", re.I)


def is_plausible_scheme(text):
    """Whether a string could be a fund's name.

    Cheap and deliberately loose: this only has to catch the page furniture
    a parser sometimes grabs, not to validate against AMFI. A name it
    rejects becomes "Scheme (name not read)" with a note, which is honest
    and fixable -- a name it wrongly accepts is a holding labelled with a
    build number, which nobody spots until they read their own portfolio.
    """
    text = (text or "").strip()
    if len(text) < 6 or FURNITURE.search(text):
        return False
    letters = sum(c.isalpha() for c in text)
    return letters >= 6 and letters >= len(text) / 3


def extract_cas_text(pdf_bytes, password=""):
    """Text of a CAS PDF, opening it with the password if it has one."""
    try:
        from pypdf import PdfReader
    except ImportError:
        raise ValueError("Reading a CAS needs the pypdf package "
                         "(pip install -r requirements.txt).")
    reader = PdfReader(io.BytesIO(pdf_bytes))
    if reader.is_encrypted:
        if not reader.decrypt(password or ""):
            raise PermissionError(
                "Wrong password. A CAS is usually locked with the password "
                "you chose when requesting it from CAMS or KFintech.")
    return "\n".join((p.extract_text() or "") for p in reader.pages)


# --------------------------------------------------------------------------
# The detailed statement
#
# One block per scheme, in this shape:
#
#     DSP Mutual Fund
#     Folio No: 4274832 / 68
#     D782-DSP Mid Cap Fund - Direct Plan - Growth (Non-Demat) -
#         ISIN: INF740K01PX1(Advisor: DIRECT)   Registrar : CAMS
#     Nominee 1: <name>   Nominee 2:   Nominee 3:
#     Opening Unit Balance: 0.000
#     Date  Transaction  Amount(INR)  Units  Price(INR)  Unit Balance
#     ... transaction rows, plus "*** Stamp Duty ***" lines ...
#     Closing Unit Balance: 7,763.079  NAV on 25-Aug-2026: INR 112.6609
#     Total Cost Value: 5,37,500.00  Market Value on 25-Aug-2026: INR 8,74,595.47
#
# One folio can carry several schemes, so blocks are cut at the ISIN and the
# folio is carried down from the header above it -- cutting at "Folio No"
# would collapse every scheme in a folio into one row.
# --------------------------------------------------------------------------
FOLIO_RE = re.compile(r"Folio\s*No[:.\s]*([0-9][\w/\- ]{2,25})", re.I)
CLOSING_RE = re.compile(r"Closing\s*Unit\s*Balance[:.\s]*([\d,]+\.?\d*)", re.I)
OPENING_RE = re.compile(r"Opening\s*Unit\s*Balance[:.\s]*([\d,]+\.?\d*)", re.I)
NAV_RE = re.compile(r"NAV\s*on\s*[\d\-A-Za-z]+[:.\s]*(?:INR|Rs\.?)?\s*"
                    r"([\d,]+\.?\d*)", re.I)
# The statements registrars send today say "Market Value on <date>"; older
# CAMS layouts say "Valuation on". Reading only the latter left every
# detailed statement with no market value at all.
VALUE_ON_RE = re.compile(r"(?:Market\s*Value|Valuation)\s*on\s*"
                         r"[\d\-A-Za-z]+[:.\s]*(?:INR|Rs\.?)?\s*"
                         r"([\d,]+\.?\d*)", re.I)
COST_RE = re.compile(r"Total\s*Cost\s*Value[:.\s]*(?:INR|Rs\.?)?\s*"
                     r"([\d,]+\.?\d*)", re.I)
# The three nominee slots print on one line, empty ones included
# ("Nominee 1: SEEMA  Nominee 2:   Nominee 3:"), so the first name has to be
# cut at the next label rather than at the end of the line.
# "HDFCFC-HDFC Flexi Cap Fund - Growth Plan". The date guard keeps a
# transaction row ("01-Jun-2024 Purchase ...") from passing as a scheme.
SCHEME_RE = re.compile(r"^\s*(?!\d{1,2}-)([A-Z0-9]{2,10})\s*[-–]\s*"
                       r"(.{6,110}?)\s*(?:\(|Registrar|ISIN|$)", re.M)
# Where one scheme's block ends and the next begins. The closing line is not
# the end of a block: the cost and market value print after it.
BLOCK_BOUNDARY = re.compile(r"Folio\s*No|Opening\s*Unit\s*Balance", re.I)
NOMINEE_RE = re.compile(r"Nominee\s*1\s*[:.]\s*([A-Za-z][A-Za-z .'’-]{2,60}?)"
                        r"\s*(?=Nominee\s*\d|$)", re.I | re.M)
DETAILED_MARKER = re.compile(r"(Opening|Closing)\s*Unit\s*Balance", re.I)

# A transaction row: date, description, amount, units, price, unit balance.
# Redemptions print in parentheses or with a minus; the stamp-duty lines
# carry no date and so cannot be mistaken for one.
TXN_RE = re.compile(
    r"^\s*(\d{2}-[A-Za-z]{3}-\d{4})\s+(.+?)\s+"
    r"\(?(-?[\d,]+\.\d{2})\)?\s+"
    r"\(?(-?[\d,]+\.\d{2,4})\)?\s+"
    r"([\d,]+\.\d{2,4})\s+"
    r"([\d,]+\.\d{2,4})\s*$")
MONTHS = {m: i + 1 for i, m in enumerate(
    ["jan", "feb", "mar", "apr", "may", "jun",
     "jul", "aug", "sep", "oct", "nov", "dec"])}


def _cas_date(token):
    """'25-Aug-2026' -> '2026-08-25'; '' when the month is not a month."""
    d, m, y = token.split("-")
    month = MONTHS.get(m.lower()[:3])
    return "%s-%02d-%02d" % (y, month, int(d)) if month else ""


def _txn_type(description):
    text = description.lower()
    if "stamp" in text or "tax" in text:
        return ""                       # a deduction, not a cashflow of yours
    if any(w in text for w in ("redemption", "switch out", "switch-out",
                               "sell", "withdrawal", "swo")):
        return "sell"
    if "idcw" in text or "dividend" in text:
        return "dividend"
    if "purchase" in text or "switch in" in text or "sip" in text \
            or "investment" in text or "systematic" in text:
        return "buy"
    return ""


def parse_cas_transactions(block):
    """Transaction rows inside one scheme block of a detailed statement.

    These are what makes a real XIRR possible, so they are read whenever the
    statement carries them; a row that does not parse cleanly is left out
    rather than guessed at, and the totals check still guards the holding.
    """
    out = []
    for line in block.splitlines():
        m = TXN_RE.match(line)
        if not m:
            continue
        iso = _cas_date(m.group(1))
        kind = _txn_type(m.group(2))
        if not iso or not kind:
            continue
        amount = to_number(m.group(3))
        units = to_number(m.group(4))
        if not amount:
            continue
        out.append({"date": iso, "type": kind,
                    "amount": round(abs(amount), 2),
                    "units": round(abs(units or 0), 4)})
    return out


def _scheme_blocks(lines):
    """(start, end) line spans, one per scheme, each holding a closing line.

    Cutting at the closing balance alone would strand the cost and market
    value that print underneath it; cutting at "Folio No" alone would
    collapse the several schemes a single folio can hold into one row. So
    each block runs to the first folio, scheme or opening-balance line
    *after* its closing line.
    """
    closings = [i for i, ln in enumerate(lines) if CLOSING_RE.search(ln)]
    spans, start = [], 0
    for c in closings:
        end = next((j for j in range(c + 1, len(lines))
                    if BLOCK_BOUNDARY.search(lines[j])
                    or ISIN_TOKEN.search(lines[j])), len(lines))
        spans.append((start, end))
        start = end
    return spans


def parse_cas(text, owner="Me"):
    """Holdings from a detailed CAS.

    Statements differ between registrars and change format, so this reports
    what it could not read instead of guessing. Every row is shown for
    confirmation before anything is imported.
    """
    lines = (text or "").splitlines()
    rows, notes = [], []
    for start, end in _scheme_blocks(lines):
        block = "\n".join(lines[start:end])
        units_m = CLOSING_RE.search(block)
        units = to_number(units_m.group(1))
        if not units or units <= 0:
            continue                      # closed folio, nothing held

        # One folio can hold several schemes and the header prints only
        # once, so take the nearest folio at or above this block's end.
        folio = ""
        for j in range(end - 1, -1, -1):
            folio_m = FOLIO_RE.search(lines[j])
            if folio_m:
                folio = folio_m.group(1).strip()
                break

        isin_m = ISIN_TOKEN.search(block)
        isin = isin_m.group(0) if isin_m else ""
        nav_m, val_m, cost_m = (NAV_RE.search(block), VALUE_ON_RE.search(block),
                                COST_RE.search(block))
        nav = to_number(nav_m.group(1)) if nav_m else None
        val = to_number(val_m.group(1)) if val_m else None
        cost = to_number(cost_m.group(1)) if cost_m else None
        if not nav and val:
            nav = val / units

        # The name sits left of the ISIN on the scheme line. Older layouts
        # print the ISIN on a line of its own, so fall back to the
        # "CODE-Scheme Name" line the block always carries.
        scheme = ""
        if isin_m:
            line = next((ln for ln in block.splitlines()
                         if ISIN_TOKEN.search(ln)), "")
            scheme = _clean_scheme(line[:ISIN_TOKEN.search(line).start()])
        if not scheme:
            scheme_m = SCHEME_RE.search(block)
            scheme = _clean_scheme(scheme_m.group(2)) if scheme_m else ""
        if not is_plausible_scheme(scheme):
            scheme = "Scheme (name not read)"
            notes.append("A scheme name could not be read for folio %s — set "
                         "it by hand after importing." % (folio or isin))
        # Transactions are only worth keeping when the statement covers the
        # whole life of the holding. An opening balance above zero means it
        # starts mid-history, and an XIRR from a truncated set of flows would
        # be confidently wrong rather than merely missing.
        txns = parse_cas_transactions(block)
        opening_m = OPENING_RE.search(block)
        opening = to_number(opening_m.group(1)) if opening_m else None
        if txns and opening:
            txns = []
            notes.append("The statement starts after you first invested in "
                         "some schemes (they open with a balance), so their "
                         "transaction history is incomplete and was not "
                         "imported. Request a CAS from an earlier date if you "
                         "want XIRR on those.")
        nominee_m = NOMINEE_RE.search(block)
        registrar_m = REGISTRAR_TOKEN.search(block)
        date_m = NAV_DATE_TOKEN.search(block[units_m.end():])
        rows.append({
            "owner": owner, "asset_class": "mutual_fund",
            "name": scheme[:120], "identifier": folio[:60], "isin": isin,
            "units": round(units, 4),
            "avg_cost": round(cost / units, 4) if cost and units else 0.0,
            "last_price": round(nav or 0.0, 4),
            "invested": round(cost or 0.0, 2),
            "current_value": round(val or (units * (nav or 0)), 2),
            "nav_date": date_m.group(0) if date_m else "",
            "registrar": registrar_m.group(1).upper() if registrar_m else "",
            "nominee": nominee_m.group(1).strip() if nominee_m else "",
            "transactions": txns,
            "purchase_date": "",
        })
    if not rows:
        notes.append("No schemes with a closing balance were found — the "
                     "format may have changed. The broker CSV route still "
                     "works.")
    return rows, sorted(set(notes))

# --------------------------------------------------------------------------
# The Consolidated Account Summary -- the table format
#
# Columns: Folio No. | ISIN | Scheme Name | Cost Value | Unit Balance |
#          NAV Date | NAV | Market Value | Registrar
#
# Extracted text wraps scheme names over several lines and often runs the
# folio straight into the ISIN, so rows are anchored on the ISIN -- exactly
# one per holding, and unmistakable -- rather than on line breaks. Every
# money column carries decimals while the digits inside scheme names
# ("NASDAQ 100", "Nifty 50", the registrar's "128TSDGG" prefix) do not, which
# is what makes the four numbers at the end of a row safe to read positionally.
# --------------------------------------------------------------------------


def parse_cas_summary(text, owner="Me"):
    """Holdings from a Consolidated Account Summary.

    Returns (rows, notes). Anything that cannot be read is reported rather
    than guessed at, and the caller confirms every row before import.
    """
    lines = [ln for ln in (text or "").splitlines() if ln.strip()]
    anchors = [i for i, ln in enumerate(lines) if ISIN_TOKEN.search(ln)]
    rows, notes = [], []
    for n, start in enumerate(anchors):
        end = anchors[n + 1] if n + 1 < len(anchors) else len(lines)
        block_text = " ".join(lines[start:end])
        isin_m = ISIN_TOKEN.search(block_text)
        isin = isin_m.group(0)
        before, after = block_text[:isin_m.start()], block_text[isin_m.end():]

        # Folio is whatever sits left of the ISIN, often glued to it.
        folio = re.sub(r"\s+", "", before).strip()
        folio = re.sub(r"^(?:Folio\s*No\.?:?)", "", folio, flags=re.I).strip()

        nums = DECIMAL_TOKEN.findall(after)
        if len(nums) < 4:
            notes.append("A row for ISIN %s had %d numeric columns, not the "
                         "expected four — check it in the preview."
                         % (isin, len(nums)))
            continue
        cost, units, nav, market = (to_number(x) for x in nums[:4])
        if not units or units <= 0:
            continue                       # exited scheme, nothing held

        scheme = _clean_scheme(after[:after.find(nums[0])])
        date_m = NAV_DATE_TOKEN.search(after)
        registrar_m = REGISTRAR_TOKEN.search(after)
        if not is_plausible_scheme(scheme):
            scheme = "Scheme (name not read)"
            notes.append("A scheme name could not be read for ISIN %s — set "
                         "it by hand after importing." % isin)
        rows.append({
            "owner": owner, "asset_class": "mutual_fund",
            "name": scheme[:120], "identifier": folio[:60], "isin": isin,
            "units": round(units, 4),
            "avg_cost": round(cost / units, 4) if cost and units else 0.0,
            "last_price": round(nav or 0.0, 4),
            "invested": round(cost or 0.0, 2),
            "current_value": round(market or (units * (nav or 0)), 2),
            "nav_date": date_m.group(0) if date_m else "",
            "registrar": registrar_m.group(1).upper() if registrar_m else "",
            "nominee": "", "transactions": [],
            "purchase_date": "",
        })
    return rows, sorted(set(notes))


TOTAL_ROW = re.compile(r"Total\s+(\d[\d,]*\.\d+)\s+(\d[\d,]*\.\d+)", re.I)


def check_against_total(text, rows):
    """Compare what was parsed against the statement's own Total row.

    A parser that quietly drops rows is worse than one that fails loudly, and
    the summary prints its own totals, so there is no excuse for not checking.
    """
    m = TOTAL_ROW.search(text or "")
    if not m:
        return []
    stated_cost, stated_market = to_number(m.group(1)), to_number(m.group(2))
    got_cost = sum(r["invested"] for r in rows)
    got_market = sum(r["current_value"] for r in rows)
    notes = []
    for label, stated, got in (("cost", stated_cost, got_cost),
                               ("market value", stated_market, got_market)):
        if stated and abs(stated - got) > max(1.0, stated * 0.005):
            notes.append(
                "The statement's total %s is %s but the rows read add up to "
                "%s — a difference of %s. Some holdings were not read; check "
                "the preview against your statement before importing."
                % (label, _fmt(stated), _fmt(got), _fmt(abs(stated - got))))
    return notes


def _fmt(x):
    return "{:,.2f}".format(x or 0)


def is_detailed_cas(text):
    """True for the transaction-by-transaction layout.

    The two layouts have to be told apart before parsing, not after: a
    detailed statement carries ISINs and plenty of decimals, so the summary
    parser will happily read a transaction row as a holding and return
    confident nonsense. The opening/closing balance lines appear only in the
    detailed layout.
    """
    return bool(DETAILED_MARKER.search(text or ""))


def parse_cas_any(text, owner="Me"):
    """Parse whichever of the two layouts arrived.

    CAMS and KFintech issue both; people request whichever they find, so the
    importer should not care which one it is handed.
    """
    if is_detailed_cas(text):
        rows, notes = parse_cas(text, owner=owner)
        if rows:
            return rows, sorted(set(notes + check_against_total(text, rows))), \
                "detailed"
    rows, notes = parse_cas_summary(text, owner=owner)
    if rows:
        return rows, sorted(set(notes + check_against_total(text, rows))), \
            "summary"
    rows, notes = parse_cas(text, owner=owner)
    if rows:
        return rows, notes, "detailed"
    return [], ["No holdings could be read. This importer understands the "
                "CAMS/KFintech Consolidated Account Summary table and the "
                "detailed statement; if yours looks different, the broker "
                "CSV route still works."], "unknown"
